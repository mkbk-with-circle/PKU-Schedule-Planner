import requests
import csv
import re
import time
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.styles import Alignment


class PKUElectiveCrawler:
    def __init__(self):
        self.session = requests.Session()
        self.base_url = "https://elective.pku.edu.cn/elective2008/edu/pku/stu/elective/controller/electiveWork/election.jsp"
        # 需要替换"Cookie"字段为有效的 Cookie"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36",
            "Cookie": "route=07f87a6dd7991fea2138ca0175e04553; _webvpn_key=eyJhbGciOiJIUzI1NiJ9.eyJ1c2VyIjoiMjIwMDAxMzE1MyIsImlhdCI6MTc2ODM3NDUzOCwiZXhwIjoxNzY4NDYwOTM4fQ.U0s2nSzAfYbDwAxTYF-AluCZp8ioS_VyqZ1rnCufgtk; webvpn_username=2200013153%7C1768374538%7C3eb6ce063ef40157c76cee769924f1c08ff3c8da; SWSPT=JgHgYSUuwTAe8t9NHuTqZvCGEWaiym39Pr5VqirQlVjQik3JreK2g2il0MSD4+jOuU0VMUqi7VGgl1gFtOQGenS730tCdHV0a7VRzT/Yp78=; JSESSIONID=jTkCp2NRGfBmhj0NnvwbgqjCk62JFLsX84ttGp9Pfnnv7fdtbKXn!480051555",
            "Referer": "https://elective.pku.edu.cn/elective2008/edu/pku/stu/elective/controller/electiveWork/ElectiveWorkController.jpf",
        }

        self.fieldnames = [
            "课程号",
            "课程名",
            "课程类别",
            "学分",
            "周学时",
            "教师",
            "班号",
            "开课单位",
            "年级",
            "上课考试信息",
            "自选PNP",
            "限数已选",
            "意愿值",
            "预选",
        ]

    def fetch_page(self, page_num=1):
        """获取指定页数的课程数据"""
        row_start = (page_num - 1) * 20

        params = {
            "netui_pagesize": "electableCourseListGrid;20",
            "netui_row": f"electableCourseListGrid;{row_start}",
        }

        try:
            response = self.session.get(
                self.base_url, headers=self.headers, params=params, timeout=10
            )
            response.encoding = "utf-8"
            return response.text
        except Exception as e:
            print(f"获取第 {page_num} 页失败: {e}")
            return None

    def parse_courses(self, html):
        """解析页面中的课程信息"""
        soup = BeautifulSoup(html, "html.parser")
        courses = []

        table = soup.find("table", class_="datagrid")
        if not table:
            print("未找到课程表格")
            return courses

        rows = table.find_all("tr", class_=re.compile(r"datagrid-(even|odd)"))

        for row in rows:
            cols = row.find_all("td")
            if len(cols) >= 14:
                course = {
                    "课程号": self.get_cell_text(cols[0]),
                    "课程名": self.get_cell_text(cols[1]),
                    "课程类别": self.get_cell_text(cols[2]),
                    "学分": self.get_cell_text(cols[3]),
                    "周学时": self.get_cell_text(cols[4]),
                    "教师": self.get_cell_text(cols[5]),
                    "班号": self.get_cell_text(cols[6]),
                    "开课单位": self.get_cell_text(cols[7]),
                    "年级": self.get_cell_text(cols[8]),
                    "上课考试信息": self.get_cell_text(cols[9]),
                    "自选PNP": self.get_cell_text(cols[10]),
                    "限数已选": self.get_cell_text(cols[11]),
                    "意愿值": self.get_cell_text(cols[12]),
                    "预选": self.get_cell_text(cols[13]),
                }
                courses.append(course)

        return courses

    def get_cell_text(self, td):
        """提取单元格文本（保留原本的换行，如 <br>）"""
        if not td:
            return ""

        # 如果是 input（比如意愿值/预选），直接取 value
        inp = td.find("input")
        if inp and inp.get("value") is not None:
            return inp.get("value", "").strip()

        # 保留结构性换行（br/div/span 等节点间用换行拼接）
        text = td.get_text(separator="\n", strip=True)

        # 清洗：去掉重复空行、统一换行
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        text = re.sub(r"\n{2,}", "\n", text)
        text = re.sub(r"[ \t]+\n", "\n", text)
        text = re.sub(r"\n[ \t]+", "\n", text)

        # 可选：如果“考试时间：”前面被挤在同一行，强制换行
        text = re.sub(r"(?<!\n)(考试时间：)", r"\n\1", text)

        return text.strip()

    def get_total_pages(self, html):
        """获取总页数"""
        soup = BeautifulSoup(html, "html.parser")
        for text in soup.stripped_strings:
            match = re.search(r"Page\s*\d+\s*of\s*(\d+)", text)
            if match:
                return int(match.group(1))
        return 1

    def get_first_course_id(self, html):
        """获取页面第一门课的课程号"""
        soup = BeautifulSoup(html, "html.parser")
        table = soup.find("table", class_="datagrid")
        if table:
            first_row = table.find("tr", class_=re.compile(r"datagrid-(even|odd)"))
            if first_row:
                first_td = first_row.find("td")
                if first_td:
                    return self.get_cell_text(first_td)
        return None

    def crawl_all(self, output_file="pku_courses.xlsx"):
        """爬取所有页面并保存到 XLSX"""
        all_courses = []

        # 自动修正扩展名
        if output_file.lower().endswith(".csv"):
            output_file = output_file[:-4] + ".xlsx"
        elif not output_file.lower().endswith(".xlsx"):
            output_file = output_file + ".xlsx"

        print("正在获取第 1 页...")
        first_page_html = self.fetch_page(1)
        if not first_page_html:
            print("获取第一页失败")
            return

        if "会话超时" in first_page_html or "重新登录" in first_page_html:
            print("❌ Cookie 已过期，请重新获取 Cookie")
            return

        total_pages = self.get_total_pages(first_page_html)
        print(f"共 {total_pages} 页")

        courses = self.parse_courses(first_page_html)
        first_page_first_id = self.get_first_course_id(first_page_html)
        all_courses.extend(courses)
        print(f"第 1 页: 获取 {len(courses)} 门课程 (首课: {first_page_first_id})")

        for page in range(2, total_pages + 1):
            time.sleep(0.5)
            print(f"正在获取第 {page} 页...")

            html = self.fetch_page(page)
            if html:
                current_first_id = self.get_first_course_id(html)
                courses = self.parse_courses(html)
                all_courses.extend(courses)
                print(
                    f"第 {page} 页: 获取 {len(courses)} 门课程 (首课: {current_first_id})"
                )

                if current_first_id == first_page_first_id:
                    print(f"⚠️ 警告: 第 {page} 页数据与第 1 页相同，翻页可能失败")

        unique_courses = self.deduplicate(all_courses)

        self.save_to_xlsx(unique_courses, output_file)
        print(
            f"\n✅ 完成！共爬取 {len(unique_courses)} 门课程（去重后），已保存到 {output_file}"
        )

    def deduplicate(self, courses):
        """根据课程号+班号去重"""
        seen = set()
        unique = []
        for course in courses:
            key = (course.get("课程号", ""), course.get("班号", ""))
            if key not in seen:
                seen.add(key)
                unique.append(course)
        return unique

    def save_to_csv(self, courses, filename):
        """（可选保留）保存课程数据到 CSV 文件"""
        if not courses:
            print("没有课程数据可保存")
            return

        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=self.fieldnames, quoting=csv.QUOTE_ALL
            )
            writer.writeheader()
            writer.writerows(courses)

    def save_to_xlsx(self, courses, filename):
        """保存课程数据到 XLSX 文件（更适合换行与中文兼容）"""
        if not courses:
            print("没有课程数据可保存")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "courses"

        # 表头
        ws.append(self.fieldnames)

        # 数据行
        for course in courses:
            ws.append([course.get(k, "") for k in self.fieldnames])

        # 让含换行的单元格自动换行显示
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):  # 跳过表头
            for cell in row:
                if isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = wrap

        # 表头对齐
        header_align = Alignment(vertical="center")
        for cell in ws[1]:
            cell.alignment = header_align

        # 简单列宽（按第一行估算，避免被多行长文本撑爆）
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = cell.value
                if v is None:
                    continue
                s = str(v).split("\n")[0]
                max_len = max(max_len, len(s))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(filename)


if __name__ == "__main__":
    crawler = PKUElectiveCrawler()
    crawler.crawl_all("pku_courses.xlsx")
