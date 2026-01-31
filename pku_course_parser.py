import csv
import re
import argparse
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# =========================
# Data Model (UI-friendly)
# =========================


class WeekPattern(str, Enum):
    EVERY = "every"  # 每周
    ODD = "odd"  # 单周
    EVEN = "even"  # 双周


WEEKDAY_MAP = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "日": 7,
    "天": 7,
}


@dataclass(frozen=True)
class CourseKey:
    """四元组唯一标识： (课程名, 课程号, 教室, 班号)"""

    course_name: str
    course_code: str
    room: str
    class_no: str


@dataclass
class Meeting:
    start_week: int
    end_week: int
    pattern: WeekPattern
    weekday: int  # 1..7
    start_period: int  # 1..12
    end_period: int  # 1..12
    room: str = ""
    raw: str = ""

    def occurs_on_week(self, week: int) -> bool:
        if week < self.start_week or week > self.end_week:
            return False
        if self.pattern == WeekPattern.EVERY:
            return True
        if self.pattern == WeekPattern.ODD:
            return week % 2 == 1
        if self.pattern == WeekPattern.EVEN:
            return week % 2 == 0
        return False

    def occurs_on(self, week: int, weekday: int, period: int) -> bool:
        return (
            self.occurs_on_week(week)
            and self.weekday == weekday
            and self.start_period <= period <= self.end_period
        )


@dataclass
class Course:
    """一行课程记录（不会丢行）。"""

    uid: Tuple[str, str]  # (课程号, 班号)
    key: CourseKey

    course_name: str
    course_code: str
    teacher: str
    department: str
    credits: float
    class_no: str = ""
    category: str = ""
    grade: str = ""
    meetings: List[Meeting] = field(default_factory=list)

    raw: Dict[str, str] = field(default_factory=dict)
    parse_warnings: List[str] = field(default_factory=list)


@dataclass
class ParseResult:
    courses: List[Course]
    by_key: Dict[CourseKey, List[Course]]
    by_uid: Dict[Tuple[str, str], Course]
    global_warnings: List[str] = field(default_factory=list)

    empty_room_rows: List[str] = field(default_factory=list)
    meeting_parse_warnings: List[str] = field(default_factory=list)
    key_collisions: List[str] = field(default_factory=list)

    total_rows: int = 0


# =========================
# Normalization helpers
# =========================

EXAM_PREFIXES = ("考试时间", "考试方式")


def _normalize_spaces(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("；", ";")
    s = s.replace("，", ",")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{2,}", "\n", s)
    return s.strip()


def _split_info_lines(info: str) -> List[str]:
    info = _normalize_spaces(info)
    if not info:
        return []
    lines = [ln.strip() for ln in info.split("\n")]
    return [ln for ln in lines if ln]


def _is_exam_line(line: str) -> bool:
    line = line.strip()
    return any(
        line.startswith(pfx + "：")
        or line.startswith(pfx + ":")
        or line.startswith(pfx)
        for pfx in EXAM_PREFIXES
    )


def _pattern_from_text(pat: Optional[str]) -> WeekPattern:
    if pat == "单周":
        return WeekPattern.ODD
    if pat == "双周":
        return WeekPattern.EVEN
    return WeekPattern.EVERY


def _strip_wrapping_parens(line: str) -> str:
    s = line.strip()
    s = s.lstrip("（(").rstrip("）)")
    s = re.sub(r"^\s*备注[:：]\s*", "", s)
    return s.strip()


def _normalize_room(room: str) -> str:
    room = room.strip()
    if not room:
        return ""
    room = re.sub(r"(机房|内)$", "", room).strip()
    room = re.sub(r"\s+", "", room)
    return room


# =========================
# Regex: 节次制 + 钟点制（备注）
# =========================

MEETING_PERIOD_RE = re.compile(
    r"""
    ^\s*
    (?P<ws>\d+)\s*~\s*(?P<we>\d+)\s*周
    \s+
    (?:(?P<pat>每周|单周|双周)\s*)?
    周(?P<wd>[一二三四五六日天])
    \s*
    (?P<ps>\d+)\s*~\s*(?P<pe>\d+)\s*节
    \s*
    (?P<room>[^（(]+?)?
    \s*(?:[（(].*)?$
    """,
    re.VERBOSE,
)

MEETING_CLOCK_RE = re.compile(
    r"""
    ^\s*
    第?\s*(?P<ws>\d+)\s*[-~～]\s*(?P<we>\d+)\s*周
    \s*周(?P<wd>[一二三四五六日天])
    \s*(?P<tod>上午|下午|晚上|晚)?\s*
    (?P<h1>\d{1,2})\s*(?:[:：](?P<m1>\d{1,2}))?\s*点?(?P<half1>半)?
    \s*[-~～]\s*
    (?P<h2>\d{1,2})\s*(?:[:：](?P<m2>\d{1,2}))?\s*点?(?P<half2>半)?
    \s*
    (?:[,，]\s*(?P<room>.+?))?
    \s*$
    """,
    re.VERBOSE,
)

CLOCK_SLOT_MAP = [
    ("下午", 13 * 60, 16 * 60 + 30, (5, 8)),
    ("晚", 18 * 60, 21 * 60 + 30, (9, 12)),
    ("晚上", 18 * 60, 21 * 60 + 30, (9, 12)),
]


def _to_minutes(h: int, m: int) -> int:
    return h * 60 + m


def _parse_clock_parts(h: str, m: Optional[str], half: Optional[str]) -> int:
    hh = int(h)
    mm = int(m) if m is not None else 0
    if half is not None:
        mm = 30
    return _to_minutes(hh, mm)


def _clock_to_periods(
    tod: Optional[str], start_min: int, end_min: int
) -> Optional[Tuple[int, int]]:
    tod_norm = (tod or "").strip()
    if tod_norm == "晚上":
        tod_norm = "晚上"
    if tod_norm == "晚":
        tod_norm = "晚"

    for label, s, e, (p1, p2) in CLOCK_SLOT_MAP:
        if tod_norm and tod_norm != label:
            continue
        if abs(start_min - s) <= 20 and abs(end_min - e) <= 20:
            return (p1, p2)

    for label, s, e, (p1, p2) in CLOCK_SLOT_MAP:
        if tod_norm and tod_norm not in (label,):
            continue
        if start_min >= s - 20 and end_min <= e + 20:
            return (p1, p2)

    if 12 * 60 <= start_min <= 15 * 60 and end_min >= 16 * 60:
        return (5, 8)
    if start_min >= 17 * 60:
        return (9, 12)
    return None


# =========================
# Room extraction (robust)
# =========================


def extract_room_from_any(line: str) -> str:
    s = _normalize_spaces(line)
    s2 = _strip_wrapping_parens(s)

    m = MEETING_PERIOD_RE.match(s)
    if m:
        room = (m.group("room") or "").strip()
        return _normalize_room(room)

    m = MEETING_CLOCK_RE.match(s2)
    if m:
        room = (m.group("room") or "").strip()
        return _normalize_room(room)

    tail = re.sub(r"[（(].*?[）)]", "", s).strip()
    tail = tail.split(",")[-1].strip()

    m2 = re.search(r"(理教|一教|二教|三教|四教)\s*([0-9]{3,4})\s*$", tail)
    if m2:
        return _normalize_room(m2.group(1) + m2.group(2))

    m3 = re.search(r"(理科\s*[一二三四五六七八九十0-9]+号楼)\s*([0-9]{3,4})", tail)
    if m3:
        return _normalize_room(m3.group(1) + m3.group(2))

    return ""


# =========================
# Meeting parsing
# =========================


def parse_meeting_line(line: str) -> Optional[Meeting]:
    raw = line.strip()
    if not raw or _is_exam_line(raw):
        return None

    s = _normalize_spaces(raw)

    m = MEETING_PERIOD_RE.match(s)
    if m:
        ws = int(m.group("ws"))
        we = int(m.group("we"))
        pat = _pattern_from_text(m.group("pat"))
        wd = WEEKDAY_MAP.get(m.group("wd"))
        if wd is None:
            return None
        ps = int(m.group("ps"))
        pe = int(m.group("pe"))
        if ps > pe:
            ps, pe = pe, ps
        room = _normalize_room((m.group("room") or "").strip())
        return Meeting(
            start_week=ws,
            end_week=we,
            pattern=pat,
            weekday=wd,
            start_period=ps,
            end_period=pe,
            room=room,
            raw=raw,
        )

    s2 = _strip_wrapping_parens(s)
    m = MEETING_CLOCK_RE.match(s2)
    if m:
        ws = int(m.group("ws"))
        we = int(m.group("we"))
        wd = WEEKDAY_MAP.get(m.group("wd"))
        if wd is None:
            return None

        tod = (m.group("tod") or "").strip()
        if tod == "晚上":
            tod_key = "晚上"
        elif tod == "晚":
            tod_key = "晚"
        else:
            tod_key = tod

        start_min = _parse_clock_parts(m.group("h1"), m.group("m1"), m.group("half1"))
        end_min = _parse_clock_parts(m.group("h2"), m.group("m2"), m.group("half2"))

        # 处理 12 小时制：下午/晚/晚上 经常写成 1-4点半、6-9点半 -> 13:00-16:30、18:00-21:30
        if tod_key in ("下午", "晚", "晚上"):
            if start_min < 12 * 60:
                start_min += 12 * 60
            if end_min < 12 * 60:
                end_min += 12 * 60

        periods = _clock_to_periods(tod_key, start_min, end_min)
        if not periods:
            return None

        p1, p2 = periods
        room = _normalize_room((m.group("room") or "").strip())
        return Meeting(
            start_week=ws,
            end_week=we,
            pattern=WeekPattern.EVERY,
            weekday=wd,
            start_period=p1,
            end_period=p2,
            room=room,
            raw=raw,
        )

    return None


# =========================
# Course row parsing
# =========================


def parse_course_row(row: Dict[str, str]) -> Course:
    course_code = str(row.get("课程号", "")).strip()
    course_name = str(row.get("课程名", "")).strip()
    teacher = str(row.get("教师", "")).strip()
    department = str(row.get("开课单位", "")).strip()
    class_no = str(row.get("班号", "")).strip()
    category = str(row.get("课程类别", "")).strip()
    grade = str(row.get("年级", "")).strip()

    credits_raw = str(row.get("学分", "")).strip()
    try:
        credits = float(credits_raw) if credits_raw != "" else 0.0
    except ValueError:
        credits = 0.0

    info = str(row.get("上课考试信息", "") or "").strip()
    lines = _split_info_lines(info)

    meetings: List[Meeting] = []
    warnings: List[str] = []
    rooms: List[str] = []

    for ln in lines:
        if _is_exam_line(ln):
            continue

        mt = parse_meeting_line(ln)
        if mt is None:
            warnings.append(f"未能解析上课行：{ln}")
            rm = extract_room_from_any(ln)
            if rm:
                rooms.append(rm)
            continue

        meetings.append(mt)
        if mt.room:
            rooms.append(mt.room)
        else:
            rm = extract_room_from_any(ln)
            if rm:
                rooms.append(rm)

    room_for_key = rooms[0] if rooms else "（地点未知）"

    key = CourseKey(
        course_name=course_name,
        course_code=course_code,
        room=room_for_key,
        class_no=class_no,
    )

    uid = (course_code, class_no)

    return Course(
        uid=uid,
        key=key,
        course_name=course_name,
        course_code=course_code,
        teacher=teacher,
        department=department,
        credits=credits,
        class_no=class_no,
        category=category,
        grade=grade,
        meetings=meetings,
        raw=row,
        parse_warnings=warnings,
    )


# =========================
# Loaders (.xlsx preferred, .csv supported)
# =========================


def load_rows_from_csv(path: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({k: (v if v is not None else "") for k, v in r.items()})
    return rows


def load_rows_from_xlsx(path: str, sheet_name: str = "courses") -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return []

    header = [str(h).strip() if h is not None else "" for h in header]
    rows: List[Dict[str, str]] = []

    for tup in it:
        if tup is None or all(v is None or str(v).strip() == "" for v in tup):
            continue
        row = {}
        for i, key in enumerate(header):
            if key == "":
                continue
            v = tup[i] if i < len(tup) else None
            row[key] = "" if v is None else str(v)
        rows.append(row)

    return rows


# =========================
# Public API (for UI stage)
# =========================


def load_courses(
    path: str, sheet_name: str = "courses", debug: bool = False
) -> ParseResult:
    """
    ✅ 第二阶段 UI 直接调用这个函数即可。
    - 返回 ParseResult（包含 courses/by_key/by_uid 等）
    - debug=False 时不打印
    """
    return parse_file(path, sheet_name=sheet_name, debug=debug)


# =========================
# Main parse entry + debug
# =========================


def parse_file(
    path: str, sheet_name: str = "courses", debug: bool = True
) -> ParseResult:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)

    if p.suffix.lower() == ".xlsx":
        rows = load_rows_from_xlsx(str(p), sheet_name=sheet_name)
    elif p.suffix.lower() == ".csv":
        rows = load_rows_from_csv(str(p))
    else:
        raise ValueError("仅支持 .xlsx 或 .csv")

    courses: List[Course] = []
    by_key: Dict[CourseKey, List[Course]] = {}
    by_uid: Dict[Tuple[str, str], Course] = {}

    global_warnings: List[str] = []
    empty_room_rows: List[str] = []
    meeting_parse_warnings: List[str] = []
    key_collisions: List[str] = []

    for idx, row in enumerate(rows, start=2):
        try:
            c = parse_course_row(row)
            courses.append(c)

            if c.uid in by_uid:
                global_warnings.append(
                    f"第{idx}行 uid重复：{c.uid} | 旧教师={by_uid[c.uid].teacher} 新教师={c.teacher}"
                )
            by_uid[c.uid] = c

            if c.key.room == "":
                empty_room_rows.append(
                    f"第{idx}行 room为空 | 课程号={c.course_code} 课程名={c.course_name} 班号={c.class_no} 教师={c.teacher} | 上课考试信息={row.get('上课考试信息','')!r}"
                )

            for w in c.parse_warnings:
                meeting_parse_warnings.append(
                    f"第{idx}行 {c.course_code}/{c.course_name}/班{c.class_no} | {w}"
                )

            lst = by_key.setdefault(c.key, [])
            if lst:
                old = lst[0]
                key_collisions.append(
                    f"第{idx}行 key碰撞：{c.key} | "
                    f"已有(班号={old.class_no},教师={old.teacher}) 又来(班号={c.class_no},教师={c.teacher})"
                )
            lst.append(c)

        except Exception as e:
            global_warnings.append(f"第 {idx} 行解析失败：{e}")

    result = ParseResult(
        courses=courses,
        by_key=by_key,
        by_uid=by_uid,
        global_warnings=global_warnings,
        empty_room_rows=empty_room_rows,
        meeting_parse_warnings=meeting_parse_warnings,
        key_collisions=key_collisions,
        total_rows=len(rows),
    )

    if debug:
        print(f"读到原始记录行数（不含表头）：{result.total_rows}")
        print(f"Course 行数（保留每一行，不丢）：{len(result.courses)}")
        print(f"CourseKey 唯一数量：{len(result.by_key)}")
        print(f"key 碰撞数量（理论上应为0，实际）：{len(result.key_collisions)}")
        print(f"room 为空的行数：{len(result.empty_room_rows)}")
        print(f"上课行未解析 warning 数：{len(result.meeting_parse_warnings)}")
        print(f"行级解析失败数：{len(result.global_warnings)}")

        if result.key_collisions:
            print("\n[示例] key碰撞（前10条）：")
            for s in result.key_collisions[:10]:
                print(" -", s)

        if result.empty_room_rows:
            print("\n[示例] room为空（前5条）：")
            for s in result.empty_room_rows[:5]:
                print(" -", s)

        if result.meeting_parse_warnings:
            print("\n[示例] 上课行未解析（前10条）：")
            for s in result.meeting_parse_warnings[:10]:
                print(" -", s)

        if result.global_warnings:
            print("\n[示例] 行级解析失败（前10条）：")
            for s in result.global_warnings[:10]:
                print(" -", s)

    return result


# =========================
# CLI entry (standalone debug)
# =========================


def _cli() -> int:
    parser = argparse.ArgumentParser(
        description="PKU 课程表解析器（可独立运行，也可被 UI import 调用）"
    )
    parser.add_argument(
        "--file", "-f", default="pku_courses.xlsx", help="输入文件路径（.xlsx/.csv）"
    )
    parser.add_argument(
        "--sheet", "-s", default="courses", help="xlsx 的 sheet 名称（默认 courses）"
    )
    parser.add_argument("--debug", action="store_true", help="打印调试信息")
    args = parser.parse_args()

    res = parse_file(args.file, sheet_name=args.sheet, debug=args.debug)

    # 保留你原来的 demo：找一条钟点制示例（如果存在）
    target = None
    for c in res.courses:
        info = c.raw.get("上课考试信息", "")
        if (
            "第" in info
            and "周" in info
            and ("下午" in info or "晚" in info or "晚上" in info)
        ):
            target = c
            break

    if args.debug and target:
        print("\n--- 钟点制示例课程 ---")
        print("uid:", target.uid)
        print("key:", target.key)
        print(
            "教师:",
            target.teacher,
            "学分:",
            target.credits,
            "开课单位:",
            target.department,
        )
        print("meetings:")
        for m in target.meetings:
            print(" -", m)
        if target.parse_warnings:
            print("warnings:")
            for w in target.parse_warnings:
                print(" -", w)

    return 0


if __name__ == "__main__":
    raise SystemExit(_cli())
